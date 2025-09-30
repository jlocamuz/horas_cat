"""
Generador de archivos Excel usando pandas
"""

import os
from datetime import datetime
from typing import Dict
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from config.default_config import DEFAULT_CONFIG


class ExcelReportGenerator:
    def __init__(self):
        self.output_dir = os.path.expanduser(DEFAULT_CONFIG['output_directory'])
        self.filename_format = DEFAULT_CONFIG['filename_format']

        # Estilos
        self.header_font = Font(bold=True, color="FFFFFF")
        self.header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        self.regular_fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
        self.extra_50_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
        self.extra_100_fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
        self.night_fill = PatternFill(start_color="D1ECF1", end_color="D1ECF1", fill_type="solid")
        self.holiday_fill = PatternFill(start_color="D6EAF8", end_color="D6EAF8", fill_type="solid")
        self.pending_fill = PatternFill(start_color="F5C6CB", end_color="F5C6CB", fill_type="solid")
        
        # Colores por turno
        self.turno_manana_fill = PatternFill(start_color="FFF9E6", end_color="FFF9E6", fill_type="solid")  # Amarillo suave
        self.turno_tarde_fill = PatternFill(start_color="FFE6F0", end_color="FFE6F0", fill_type="solid")   # Rosa suave
        self.turno_noche_fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")   # Lavanda suave

        self.thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        self.center_alignment = Alignment(horizontal='center', vertical='center')

    # -------- público --------

    def generate_report(self, processed_data: Dict, start_date: str, end_date: str, output_filename: str = None) -> str:
        """Genera el reporte Excel completo usando pandas."""
        if not output_filename:
            output_filename = self.filename_format.format(
                start_date=start_date.replace('-', ''), 
                end_date=end_date.replace('-', '')
            )
        
        os.makedirs(self.output_dir, exist_ok=True)
        filepath = os.path.join(self.output_dir, output_filename)

        # Crear DataFrames
        df_summary = self._create_summary_dataframe(processed_data)
        df_daily = self._create_daily_dataframe(processed_data)

        # Escribir a Excel con pandas
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            df_summary.to_excel(writer, sheet_name='Resumen Consolidado', index=False, startrow=4)
            df_daily.to_excel(writer, sheet_name='Detalle Diario', index=False, startrow=4)

        # Aplicar estilos con openpyxl
        self._apply_styles(filepath, start_date, end_date, df_summary, df_daily)
        
        print(f"✅ Reporte Excel generado: {filepath}")
        return filepath

    # -------- DataFrames --------

    def _create_summary_dataframe(self, processed_data: Dict) -> pd.DataFrame:
        """Crea DataFrame del resumen consolidado."""
        rows = []
        for emp in processed_data.values():
            info = emp['employee_info']
            totals = emp['totals']
            rows.append({
                'ID Empleado': info.get('employeeInternalId', ''),
                'Nombre': info.get('firstName', ''),
                'Apellido': info.get('lastName', ''),
                'Turno': info.get('turno', ''),
                'Total Horas': round(totals.get('total_hours_worked', 0.0), 2),
                'Horas Regulares': round(totals.get('total_regular_hours', 0.0), 2),
                'Horas Extra 50%': round(totals.get('total_extra_hours_50', 0.0), 2),
                'Horas Extra 100%': round(totals.get('total_extra_hours_100', 0.0), 2),
                'Horas Nocturnas': round(totals.get('total_night_hours', 0.0), 2),
                'Horas Feriado': round(totals.get('total_holiday_hours', 0.0), 2),
            })
        
        df = pd.DataFrame(rows)
        
        # Ordenar por turno: Mañana -> Tarde -> Noche -> Otros/Vacío
        turno_order = {'Mañana': 1, 'Tarde': 2, 'Noche': 3}
        df['turno_sort'] = df['Turno'].map(lambda x: turno_order.get(x, 999))
        df = df.sort_values('turno_sort').drop('turno_sort', axis=1)
        df = df.reset_index(drop=True)
        
        return df

    def _create_daily_dataframe(self, processed_data: Dict) -> pd.DataFrame:
        """Crea DataFrame del detalle diario."""
        rows = []
        for emp in processed_data.values():
            info = emp['employee_info']
            for d in emp['daily_data']:
                observations = []
                if d.get('is_holiday'):
                    observations.append(f"Feriado: {d.get('holiday_name') or 'N/A'}")
                if d.get('has_time_off'):
                    observations.append(f"Licencia: {d.get('time_off_name') or 'N/A'}")
                if d.get('has_absence'):
                    observations.append("Ausencia")
                if d.get('pending_hours', 0) > 0:
                    observations.append(f"{d['pending_hours']:.1f}h pendientes")
                if d.get('day_of_week') in ['Sábado', 'Domingo']:
                    observations.append("Fin de semana")

                rows.append({
                    'ID Empleado': info.get('employeeInternalId', ''),
                    'Nombre': info.get('firstName', ''),
                    'Apellido': info.get('lastName', ''),
                    'Fecha': d.get('date', ''),
                    'Día': d.get('day_of_week', ''),
                    'Turno': d.get('turno', ''),
                    'Inicio Turno': d.get('shift_start', ''),
                    'Fin Turno': d.get('shift_end', ''),
                    'Es Franco': 'Sí' if d.get('is_rest_day') else 'No',
                    'Horas Trabajadas': round(d.get('hours_worked', 0.0), 2),
                    'Horas Regulares': round(d.get('regular_hours', 0.0), 2),
                    'Horas Extra 50%': round(d.get('extra_hours_50', 0.0), 2),
                    'Horas Extra 100%': round(d.get('extra_hours_100', 0.0), 2),
                    'Horas Nocturnas': round(d.get('night_hours', 0.0), 2),
                    'Horas Feriado': round(d.get('holiday_hours', 0.0), 2),
                    'Horas Pendientes': round(d.get('pending_hours', 0.0), 2),
                    'Es Feriado': 'Sí' if d.get('is_holiday') else 'No',
                    'Nombre Feriado': d.get('holiday_name') or '',
                    'Tiene Licencia': 'Sí' if d.get('has_time_off') else 'No',
                    'Tipo Licencia': d.get('time_off_name') or '',
                    'Tiene Ausencia': 'Sí' if d.get('has_absence') else 'No',
                    'Observaciones': ', '.join(observations) if observations else '',
                    'Cálculo (explicación)': d.get('calc_note', '')
                })
        return pd.DataFrame(rows)

    # -------- Estilos --------

    def _apply_styles(self, filepath: str, start_date: str, end_date: str, 
                      df_summary: pd.DataFrame, df_daily: pd.DataFrame):
        """Aplica estilos a las hojas del Excel."""
        wb = load_workbook(filepath)
        
        self._style_summary_sheet(wb['Resumen Consolidado'], start_date, end_date, df_summary)
        self._style_daily_sheet(wb['Detalle Diario'], start_date, end_date, df_daily)
        
        wb.save(filepath)

    def _style_summary_sheet(self, ws, start_date: str, end_date: str, df: pd.DataFrame):
        """Aplica estilos a la hoja de resumen."""
        # Títulos
        ws['A1'] = "REPORTE DE ASISTENCIA - RESUMEN CONSOLIDADO"
        ws['A2'] = f"Período: {start_date} al {end_date}"
        ws['A3'] = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        for row in range(1, 4):
            ws[f'A{row}'].font = Font(bold=True, size=12)

        # Headers (fila 5)
        for col in range(1, len(df.columns) + 1):
            cell = ws.cell(row=5, column=col)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.thin_border
            cell.alignment = self.center_alignment

        # Datos, colores y colores por turno
        for row_idx in range(len(df)):
            row = row_idx + 6
            turno = df.iloc[row_idx]['Turno']
            
            # Determinar color de fondo por turno
            turno_fill = None
            if turno == 'Mañana':
                turno_fill = self.turno_manana_fill
            elif turno == 'Tarde':
                turno_fill = self.turno_tarde_fill
            elif turno == 'Noche':
                turno_fill = self.turno_noche_fill
            
            for col in range(1, len(df.columns) + 1):
                cell = ws.cell(row=row, column=col)
                cell.border = self.thin_border
                
                # Aplicar solo color de turno a toda la fila
                if turno_fill:
                    cell.fill = turno_fill

        # Totales
        total_row = 6 + len(df) + 1
        ws.cell(row=total_row, column=1, value="TOTALES").font = Font(bold=True)
        for col in range(5, 11):  # Total Horas hasta Horas Feriado
            col_letter = get_column_letter(col)
            cell = ws.cell(row=total_row, column=col, value=f"=SUM({col_letter}6:{col_letter}{5 + len(df)})")
            cell.font = Font(bold=True)
            cell.border = self.thin_border

        # Anchos de columna
        for col in range(1, len(df.columns) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 17

    def _style_daily_sheet(self, ws, start_date: str, end_date: str, df: pd.DataFrame):
        """Aplica estilos a la hoja de detalle diario."""
        # Títulos
        ws['A1'] = "DETALLE DIARIO DE ASISTENCIA"
        ws['A2'] = f"Período: {start_date} al {end_date}"
        ws['A3'] = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        for row in range(1, 4):
            ws[f'A{row}'].font = Font(bold=True, size=12)

        # Headers
        for col in range(1, len(df.columns) + 1):
            cell = ws.cell(row=5, column=col)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.thin_border
            cell.alignment = self.center_alignment

        # Datos y colores
        for row_idx in range(len(df)):
            row = row_idx + 6
            for col in range(1, len(df.columns) + 1):
                cell = ws.cell(row=row, column=col)
                cell.border = self.thin_border
                
                # Colores por métricas
                if col == 11: cell.fill = self.regular_fill     # Horas Regulares
                elif col == 12: cell.fill = self.extra_50_fill  # Extra 50
                elif col == 13: cell.fill = self.extra_100_fill # Extra 100
                elif col == 14: cell.fill = self.night_fill     # Nocturnas
                elif col == 15: cell.fill = self.holiday_fill   # Horas Feriado
                elif col == 16: cell.fill = self.pending_fill   # Pendientes

        # Anchos de columna
        col_widths = {22: 28, 23: 55, 18: 22}  # Observaciones, Explicación, Nombre Feriado
        for col in range(1, len(df.columns) + 1):
            width = col_widths.get(col, 14)
            ws.column_dimensions[get_column_letter(col)].width = width

        # Wrap text para observaciones y explicación
        for row in range(6, 6 + len(df)):
            ws.cell(row=row, column=22).alignment = Alignment(wrap_text=True, vertical='top')
            ws.cell(row=row, column=23).alignment = Alignment(wrap_text=True, vertical='top')